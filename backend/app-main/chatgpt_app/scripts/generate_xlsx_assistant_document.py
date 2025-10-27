import os
import tempfile
import openpyxl
from openpyxl.utils import get_column_letter

def run_generate_xlsx(query_result):
    # Extract the active_directory_query_result
    data = query_result.get('active_directory_query_result', [])
    if not data:
        raise ValueError("No data found in active_directory_query_result")

    # Flatten the data
    flattened_data = []
    for entry in data:
        flattened_entry = {}
        for key, value in entry.items():
            if isinstance(value, list):
                # Join list items with a comma
                flattened_entry[key] = ', '.join(map(str, value))
            else:
                flattened_entry[key] = value
        flattened_data.append(flattened_entry)

    # Create a workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write headers
    if flattened_data:
        headers = flattened_data[0].keys()
        ws.append(list(headers))

        # Write data rows
        for row_data in flattened_data:
            row = [row_data.get(header, '') for header in headers]
            ws.append(row)
    else:
        # No data, write "No data available"
        ws.append(["No data available"])

    # Adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    # Save the workbook to a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    return tmp_path  # Return the path to the temporary file
